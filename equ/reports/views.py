from django.http import HttpResponse
from django.shortcuts import render
from equipments.models import Equipment, Printer
from django.core.paginator import Paginator
import pandas as pd
from datetime import datetime

from inventory.models import Inventory
from .models import Report
from django.core.files.base import File
from io import BytesIO
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.mixins import LoginRequiredMixin
from .mixin import AccountingUserRequiredMixin
from django.views.generic import ListView

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Report


def export_excel(request):
    # Создание книги
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Отчет по оборудованию'

    # Добавление заголовков
    headers = ['ID', 'Наименование', 'Местоположение', 'Отв. сотрудник', 'Тип']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    # Добавление данных
    equipments = Equipment.objects.all()  # Замените на ваш QuerySet
    for row_num, equipment in enumerate(equipments, 2):
        sheet[f'A{row_num}'] = equipment.pk
        sheet[f'B{row_num}'] = equipment.title
        # Преобразование location в строку (например, название или другое поле)
        sheet[f'C{row_num}'] = str(equipment.location) if equipment.location else ''
        # Преобразование ответственного сотрудника
        sheet[f'D{row_num}'] = (
            f"{equipment.responsible.first_name} {equipment.responsible.last_name}" 
            if equipment.responsible else ''
        )
        # Преобразование категории
        sheet[f'E{row_num}'] = str(equipment.category) if equipment.category else ''

    # Ответ с Excel-файлом
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=equipment_report.xlsx'
    workbook.save(response)
    return response

# Create your views here.


# Класс для списка оборудования

class EquipmentReportView(LoginRequiredMixin, AccountingUserRequiredMixin, ListView):
    model = Equipment
    template_name = 'reports/report_equipments.html'
    context_object_name = 'objects'  # Соответствует вашему шаблону
    paginate_by = 10  # Количество объектов на странице

    def get(self, request, *args, **kwargs):
        # Получаем все объекты Equipment
        equipments = Equipment.objects.all()

        # Создаем объект Paginator для пагинации
        paginator = Paginator(equipments, self.paginate_by)

        # Получаем номер страницы из GET-параметров
        page_number = request.GET.get('page')

        # Получаем текущую страницу с данными
        page_obj = paginator.get_page(page_number)

        # Передаем контекст в шаблон
        context = {
            'objects': page_obj
        }
        return render(request, self.template_name, context)


# Класс для списка принтеров
class PrinterReportView(LoginRequiredMixin, AccountingUserRequiredMixin, ListView):
    model = Printer
    template_name = 'reports/report_printers.html'
    context_object_name = 'objects'


# Класс для списка инвентарей
class InventoryReportView(LoginRequiredMixin, AccountingUserRequiredMixin, ListView):
    model = Inventory
    template_name = 'reports/report_inventorys.html'
    context_object_name = 'objects'
    paginate_by = 5

    def get(self, request, *args, **kwargs):
        inventorys = Inventory.objects.all()
        paginator = Paginator(inventorys, self.paginate_by)

        # Получаем номер страницы из GET-параметров
        page_number = request.GET.get('page')

        # Получаем текущую страницу с данными
        page_obj = paginator.get_page(page_number)

        # Передаем контекст в шаблон
        context = {
            'objects': page_obj
        }
        return render(request, self.template_name, context)
